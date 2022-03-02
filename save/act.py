import os.path
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from extract import headers, base_url
from extract.act import get_act_details, get_links, get_txt
from extract.detector import detect_type
from save import files_dir, files_list_dir, txt_files_dir, txt_files_dir_converted
from utils import fix_dir_name, trim, convert_xht_to_txt_2

excel = pd.read_excel(files_list_dir,)
wb = load_workbook(files_list_dir)
ws = wb.worksheets[0]

# wb2 = load_workbook(ref_list_dir)
# ws2 = wb2.worksheets[0]


def dl_file(url,name,_dir):
    path = files_dir+"/"+_dir+"/"+name
    if os.path.isfile(files_dir+"/"+name):
        print(f'file {url} already exist as {path} .skipping...')
        return
    try:
        r = requests.get(url)
        open(path, 'wb').write(r.content)
    except:
        return None
    return True


def already_added(title_):
    result_count = len(excel.loc[excel['title'] == title_])
    return result_count > 0
    # t_index = -1
    # y_index = -1
    # n_index = -1
    # for index , t in enumerate(excel['type']):
    #     if str(t)==type_:
    #         t_index = index
    # for index , t in enumerate(excel['year']):
    #     if str(t)==year:
    #         y_index = index
    # for index , t in enumerate(excel['number']):
    #     if str(t)==num:
    #         n_index = index
    # if t_index == y_index and t_index==n_index and t_index != -1:
    #     return True
    # else:
    #     return False



def append_act(p_id:dict):
    # temp = p_id['pid'].split("/")
    # type_ = temp[0]
    # year = p_id['year']
    # num = p_id['number']
    # act_detail = get_act_details(p_id['pid'])
    title = fix_dir_name(p_id['title'].strip())
    # if already_added(trim(title)):
    #     print(f'Act {p_id["pid"]} already loaded.')
    #     return True
    # type_ = detect_type(type_,title)
    # if type_ is None:
    #     print(f'Act {p_id["pid"]} is not  included in accepted types.')
    #     return None
    # print({p_id["pid"]})
    # downloaded = None
    # for key in act_detail['files'].keys():
        # status = dl_file(act_detail['files'][key],trim(title)+key,key.replace(".",""))
        # if status is None:
        #     continue
        # if "xht" in key:
            # print(act_detail['files'][key])
            # txt = get_txt(act_detail['files'][key])
            # text = convert_xht_to_txt_2(str(txt))
            # if len(text) == 0:
            #     continue
            # f = open(txt_files_dir+"/"+trim(title) + key[:-3] + "txt", "w")
            # for line in text:
            #     f.write(line)
            # f.close()
            # downloaded = True

    # if downloaded:
    #     ws.append([trim(type_), trim(year), trim(num), trim(title), trim(act_detail['extend']), trim(act_detail['note'])])
    #     wb.save(files_list_dir)

    txt = get_txt(base_url+p_id['url'])
    text = convert_xht_to_txt_2(str(txt))
    if len(text) == 0:
        return False
    f = open(txt_files_dir+"/"+trim(title) + ".txt", "w")
    for line in text:
        f.write(line)
    f.close()

    return True



# append_act("ukpga/Edw7/6/14")
# append_act("uksi/2015/1561")

# def add_links(p_id,link):
#     ws2.append([p_id, link])
#     wb2.save(ref_list_dir)










# append_act("eudn/2020/2255")